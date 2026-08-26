import io
import re
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any, Tuple, Optional

CNPJ_REGEX = re.compile(r'\d{11,14}')

def find_cnpj_and_name_columns(df: pd.DataFrame) -> Tuple[Optional[str], Optional[str]]:
    cnpj_col = None
    name_col = None

    # Step 1: Search by Header Names
    for col in df.columns:
        col_lower = str(col).lower().strip()
        if not cnpj_col and any(k in col_lower for k in ["cnpj", "cpf/cnpj", "doc", "documento", "c.n.p.j"]):
            cnpj_col = col
        elif not name_col and any(k in col_lower for k in ["empresa", "nome", "razao", "razão", "cliente", "fantasia", "denominacao"]):
            name_col = col

    # Step 2: Search by Cell Content
    if not cnpj_col:
        for col in df.columns:
            # Check first 20 non-null values
            sample_vals = df[col].dropna().astype(str).head(20)
            digit_matches = sum(1 for val in sample_vals if len("".join(filter(str.isdigit, val))) in [11, 12, 13, 14])
            if sample_vals.shape[0] > 0 and (digit_matches / sample_vals.shape[0]) >= 0.5:
                cnpj_col = col
                break

    if not name_col:
        for col in df.columns:
            if col != cnpj_col:
                sample_vals = df[col].dropna().astype(str).head(20)
                text_matches = sum(1 for val in sample_vals if any(c.isalpha() for c in val))
                if sample_vals.shape[0] > 0 and (text_matches / sample_vals.shape[0]) >= 0.5:
                    name_col = col
                    break

    # Fallback to first two columns if detection fails
    if not cnpj_col and len(df.columns) > 0:
        cnpj_col = df.columns[0]
    if not name_col and len(df.columns) > 1:
        name_col = df.columns[1] if df.columns[1] != cnpj_col else (df.columns[0] if len(df.columns) == 1 else None)

    return cnpj_col, name_col

def parse_excel_or_csv(file_bytes: bytes, filename: str) -> List[Dict[str, str]]:
    filename_lower = filename.lower()
    
    if filename_lower.endswith(".csv"):
        try:
            df = pd.read_csv(io.BytesIO(file_bytes), dtype=str)
        except Exception:
            df = pd.read_csv(io.BytesIO(file_bytes), dtype=str, sep=";")
    else:
        df = pd.read_excel(io.BytesIO(file_bytes), dtype=str, header=None if pd.read_excel(io.BytesIO(file_bytes)).shape[0] < 2 else 0)

    # Clean empty rows and columns
    df = df.dropna(how="all").dropna(axis=1, how="all")

    cnpj_col, name_col = find_cnpj_and_name_columns(df)

    items = []
    seen = set()

    for idx, row in df.iterrows():
        raw_cnpj = str(row[cnpj_col]) if cnpj_col and pd.notna(row[cnpj_col]) else ""
        raw_name = str(row[name_col]) if name_col and pd.notna(row[name_col]) else ""

        # Extract digits
        digits = "".join(filter(str.isdigit, raw_cnpj))
        
        # Ignore header row if digits is not a valid CNPJ length or is header label
        if not digits or len(digits) < 8 or digits.isalpha():
            continue

        clean_cnpj = digits.zfill(14)
        
        if clean_cnpj not in seen and len(clean_cnpj) == 14:
            seen.add(clean_cnpj)
            items.append({
                "cnpj": clean_cnpj,
                "nome": raw_name.strip()
            })

    return items

def format_cnpj_mask(cnpj: str) -> str:
    c = "".join(filter(str.isdigit, str(cnpj))).zfill(14)
    if len(c) == 14:
        return f"{c[:2]}.{c[2:5]}.{c[5:8]}/{c[8:12]}-{c[12:]}"
    return cnpj

def generate_export_excel(results: List[Dict[str, Any]]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumo Regime Tributário"

    # Ensure grid lines are visible
    ws.views.sheetView[0].showGridLines = True

    # Title Banner
    ws.merge_cells("A1:O1")
    title_cell = ws["A1"]
    title_cell.value = "RELATÓRIO DE CONSULTA DE REGIME TRIBUTÁRIO (SIMPLES NACIONAL & MEI)"
    title_cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Subtitle with date
    ws.merge_cells("A2:O2")
    sub_cell = ws["A2"]
    sub_cell.value = f"Gerado via Base Aberta Receita Federal em {pd.Timestamp.now().strftime('%d/%m/%Y às %H:%M')}"
    sub_cell.font = Font(name="Calibri", size=10, italic=True, color="4B5563")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    headers = [
        "CNPJ",
        "Empresa (Planilha)",
        "Razão Social (Receita)",
        "Nome Fantasia",
        "Regime Tributário",
        "Optante Simples?",
        "Data Opção Simples",
        "Data Exclusão Simples",
        "Optante MEI?",
        "Data Opção MEI",
        "Data Exclusão MEI",
        "Situação Cadastral",
        "CNAE Principal",
        "UF / Município",
        "Data Consulta"
    ]

    header_row_idx = 4
    ws.row_dimensions[header_row_idx].height = 25

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=header_row_idx, column=col_num)
        cell.value = header_title
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Fills for rows
    fill_simples = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid") # Soft Green
    fill_mei = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")     # Soft Indigo/Blue
    fill_excluido = PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid")# Soft Orange
    fill_outro = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")   # Soft Gray
    fill_baixada = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") # Soft Red

    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )

    row_start = 5
    for idx, item in enumerate(results):
        curr_row = row_start + idx
        ws.row_dimensions[curr_row].height = 22

        resumo = item.get("regime_resumo", "")
        opcao_simples = "SIM" if item.get("opcao_pelo_simples") is True else ("NÃO" if item.get("opcao_pelo_simples") is False else "-")
        opcao_mei = "SIM" if item.get("opcao_pelo_mei") is True else ("NÃO" if item.get("opcao_pelo_mei") is False else "-")
        
        uf_mun = f"{item.get('uf', '')} / {item.get('municipio', '')}".strip(" /")

        row_data = [
            format_cnpj_mask(item.get("cnpj", "")),
            item.get("nome_planilha", "") or item.get("razao_social", ""),
            item.get("razao_social", ""),
            item.get("nome_fantasia", ""),
            resumo,
            opcao_simples,
            item.get("data_opcao_simples", "") or "-",
            item.get("data_exclusao_simples", "") or "-",
            opcao_mei,
            item.get("data_opcao_mei", "") or "-",
            item.get("data_exclusao_mei", "") or "-",
            item.get("situacao_cadastral", ""),
            item.get("cnae_descricao", ""),
            uf_mun,
            item.get("consulted_at", "")
        ]

        # Determine row highlight based on regime
        row_fill = fill_outro
        if "MEI" in resumo:
            row_fill = fill_mei
        elif "Simples Nacional" in resumo:
            row_fill = fill_simples
        elif "Excluído" in resumo:
            row_fill = fill_excluido
        elif "Situação Cadastral" in resumo or "INVÁLIDO" in item.get("situacao_cadastral", "").upper():
            row_fill = fill_baixada

        for col_num, val in enumerate(row_data, 1):
            cell = ws.cell(row=curr_row, column=col_num)
            cell.value = val
            cell.border = thin_border
            cell.fill = row_fill
            cell.font = Font(name="Calibri", size=10)
            
            # Alignments
            if col_num in [1, 6, 7, 8, 9, 10, 11, 12, 15]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row < 4:
                continue
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
